# encoding=utf-8
'''
将获取的数据转化为EXCEL
'''
import openpyxl


def get_excel(data, field, file):
    # 新建工作簿对象
    work = openpyxl.Workbook()
    ##激活一个sheet
    sheet = work.active
    # 给sheet命名
    sheet.title = '数据展示'
    # 将数据和字段写入excel函数
    for col in range(len(field)):
        # row：行数 column:列数 value:单元格， 注意：行列都从1开始计数
        x = sheet.cell(row=1, column=col + 1, value=u'%s' % field[col][0])
    for row in range(len(data)):
        for col in range(len(field)):
            # 因为第一行写了字段名称，所以要从第二行开始写入
            x = sheet.cell(row=row + 2, column=col + 1, value=u'%s' % data[row][col])
    # 保存excel
    newWork = work.save(file)
    return newWork
